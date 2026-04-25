"""
Aplikasi Penghitung Buah menggunakan OpenCV + Ekspor Excel
==========================================================
Cara pakai:
  python fruit_counter.py                  -> buka webcam
  python fruit_counter.py gambar.jpg       -> dari file gambar
  python fruit_counter.py gambar.jpg --save hasil.jpg  -> simpan hasil

Kontrol:
  q / ESC  -> keluar & simpan ke Excel
  s        -> simpan screenshot + catat ke Excel (mode kamera)
  +/-      -> naik/turun sensitivitas area minimum
  d        -> toggle debug mask
"""

import cv2
import numpy as np
import argparse
import sys
import os
from datetime import datetime
from dataclasses import dataclass
from typing import List, Tuple, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Konfigurasi rentang warna HSV ─────────────────────────────────────────────
FRUIT_COLORS = {
    "merah": [
        (np.array([0,  120, 70]),  np.array([10, 255, 255]), (50,  50, 220), "Apel/Semangka"),
        (np.array([170,120, 70]),  np.array([180,255, 255]), (50,  50, 220), "Apel/Semangka"),
    ],
    "oranye": [
        (np.array([10, 120, 100]), np.array([25, 255, 255]), (30, 140, 240), "Jeruk/Mangga"),
    ],
    "kuning": [
        (np.array([25, 100,  80]), np.array([35, 255, 255]), (30, 220, 240), "Pisang/Lemon"),
    ],
    "hijau": [
        (np.array([35,  80,  50]), np.array([85, 255, 255]), (60, 180,  60), "Jeruk Nipis/Alpukat"),
    ],
    "ungu": [
        (np.array([130, 80,  50]), np.array([160,255, 255]), (180, 60, 130), "Anggur/Manggis"),
    ],
}

MEAL_TIMES = {"1": "Lunch", "2": "Dinner", "3": "Supper"}

# ── Warna tema Excel ──────────────────────────────────────────────────────────
HEADER_BG  = "2E7D32"
HEADER_FG  = "FFFFFF"
SUBHDR_BG  = "A5D6A7"
ROW_ODD    = "F1F8E9"
ROW_EVEN   = "FFFFFF"
TOTAL_BG   = "C8E6C9"
BORDER_CLR = "B0BEC5"


@dataclass
class FruitDetection:
    x: int; y: int; w: int; h: int
    radius: int; cx: int; cy: int
    area: float
    color_key: str; label: str
    display_color: Tuple[int, int, int]


# ══════════════════════════════════════════════════════════════════════════════
#  Input Pengguna
# ══════════════════════════════════════════════════════════════════════════════

def ask_session_info() -> Tuple[str, str, Optional[str]]:
    fruit_keys  = list(FRUIT_COLORS.keys())
    fruit_names = [FRUIT_COLORS[k][0][3] for k in fruit_keys]

    print("\n" + "-" * 50)
    print("  Informasi Sesi Penghitungan")
    print("-" * 50)

    while True:
        client = input("  Nama Klien : ").strip()
        if client:
            break
        print("  [!] Nama klien tidak boleh kosong.")

    print("  Waktu Penyediaan:")
    print("    1. Lunch")
    print("    2. Dinner")
    print("    3. Supper")
    while True:
        choice = input("  Pilih (1/2/3) : ").strip()
        if choice in MEAL_TIMES:
            meal_time = MEAL_TIMES[choice]
            break
        print("  [!] Pilih angka 1, 2, atau 3.")

    print("  Jenis Buah yang Dihitung:")
    for i, name in enumerate(fruit_names, 1):
        print(f"    {i}. {name}")
    print(f"    {len(fruit_names) + 1}. Semua Buah")
    while True:
        max_opt = len(fruit_names) + 1
        choice  = input(f"  Pilih (1-{max_opt}) : ").strip()
        if choice == str(max_opt):
            selected_key = None
            fruit_label  = "Semua Buah"
            break
        elif choice.isdigit() and 1 <= int(choice) <= len(fruit_names):
            idx          = int(choice) - 1
            selected_key = fruit_keys[idx]
            fruit_label  = fruit_names[idx]
            break
        print(f"  [!] Pilih angka 1 sampai {max_opt}.")

    print(f"\n  OK  Klien : {client}")
    print(f"      Waktu : {meal_time}")
    print(f"      Buah  : {fruit_label}")
    print("-" * 50 + "\n")
    return client, meal_time, selected_key


# ══════════════════════════════════════════════════════════════════════════════
#  Deteksi Buah
# ══════════════════════════════════════════════════════════════════════════════

def preprocess(frame: np.ndarray) -> np.ndarray:
    return cv2.cvtColor(cv2.GaussianBlur(frame, (11, 11), 0), cv2.COLOR_BGR2HSV)


def get_mask_for_color(hsv: np.ndarray, color_key: str) -> np.ndarray:
    combined = np.zeros(hsv.shape[:2], dtype=np.uint8)
    for lower, upper, _, _ in FRUIT_COLORS[color_key]:
        combined = cv2.bitwise_or(combined, cv2.inRange(hsv, lower, upper))
    kernel   = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (9, 9))
    combined = cv2.morphologyEx(combined, cv2.MORPH_CLOSE, kernel, iterations=2)
    combined = cv2.morphologyEx(combined, cv2.MORPH_OPEN,  kernel, iterations=1)
    return cv2.dilate(combined, kernel, iterations=1)


def detect_fruits(frame: np.ndarray, min_area: int = 1500,
                  selected_key: Optional[str] = None) -> List[FruitDetection]:
    hsv        = preprocess(frame)
    detections = []
    keys       = [selected_key] if selected_key else list(FRUIT_COLORS.keys())

    for color_key in keys:
        mask        = get_mask_for_color(hsv, color_key)
        contours, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        _, _, display_color, label = FRUIT_COLORS[color_key][0]

        for cnt in contours:
            area = cv2.contourArea(cnt)
            if area < min_area:
                continue
            perimeter = cv2.arcLength(cnt, True)
            if perimeter == 0:
                continue
            if (4 * np.pi * area) / (perimeter ** 2) < 0.55:
                continue
            x, y, w, h = cv2.boundingRect(cnt)
            if min(w, h) / max(w, h) < 0.55:
                continue
            hull_area = cv2.contourArea(cv2.convexHull(cnt))
            if hull_area > 0 and (area / hull_area) < 0.75:
                continue
            (cx, cy), radius = cv2.minEnclosingCircle(cnt)
            detections.append(FruitDetection(
                x=x, y=y, w=w, h=h,
                radius=int(radius), cx=int(cx), cy=int(cy),
                area=area, color_key=color_key,
                label=label, display_color=display_color,
            ))
    return detections


# ══════════════════════════════════════════════════════════════════════════════
#  Tampilan OpenCV
# ══════════════════════════════════════════════════════════════════════════════

def draw_results(frame: np.ndarray, detections: List[FruitDetection],
                 client: str, meal_time: str, fruit_label: str) -> np.ndarray:
    output = frame.copy()
    counts = {}

    for i, det in enumerate(detections, 1):
        c = det.display_color
        cv2.circle(output, (det.cx, det.cy), det.radius, c, 2)
        cv2.circle(output, (det.cx, det.cy), 4, c, -1)
        cv2.putText(output, str(i), (det.cx - 8, det.cy + 6),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.65, (255,255,255), 2, cv2.LINE_AA)
        cv2.putText(output, str(i), (det.cx - 8, det.cy + 6),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.65, c, 1, cv2.LINE_AA)
        lbl_y = det.cy + det.radius + 18
        cv2.putText(output, det.label, (det.cx - 40, lbl_y),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.42, (255,255,255), 3, cv2.LINE_AA)
        cv2.putText(output, det.label, (det.cx - 40, lbl_y),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.42, c, 1, cv2.LINE_AA)
        counts[det.label] = counts.get(det.label, 0) + 1

    panel_h = 75 + len(counts) * 26 + 10
    overlay = output.copy()
    cv2.rectangle(overlay, (8, 8), (310, panel_h), (20,20,20), -1)
    cv2.addWeighted(overlay, 0.65, output, 0.35, 0, output)
    cv2.rectangle(output, (8, 8), (310, panel_h), (80,80,80), 1)

    cv2.putText(output, f"Klien : {client}", (16, 28),
                cv2.FONT_HERSHEY_SIMPLEX, 0.48, (200,230,255), 1, cv2.LINE_AA)
    cv2.putText(output, f"{meal_time}  |  {fruit_label}", (16, 48),
                cv2.FONT_HERSHEY_SIMPLEX, 0.44, (200,230,255), 1, cv2.LINE_AA)
    cv2.putText(output, f"Total : {len(detections)} buah", (16, 68),
                cv2.FONT_HERSHEY_SIMPLEX, 0.65, (255,255,255), 2, cv2.LINE_AA)
    y_pos = 94
    for lbl, cnt in sorted(counts.items()):
        cv2.putText(output, f"  {lbl}: {cnt}", (16, y_pos),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.48, (200,240,200), 1, cv2.LINE_AA)
        y_pos += 26

    return output


# ══════════════════════════════════════════════════════════════════════════════
#  Excel Export
# ══════════════════════════════════════════════════════════════════════════════

def _thin_border() -> Border:
    s = Side(style="thin", color=BORDER_CLR)
    return Border(left=s, right=s, top=s, bottom=s)


def _cell(ws, row, col, value, bold=False, bg=None, fg="000000", align="left"):
    cell           = ws.cell(row=row, column=col, value=value)
    cell.font      = Font(name="Arial", bold=bold, color=fg, size=11)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border    = _thin_border()
    if bg:
        cell.fill = PatternFill("solid", start_color=bg)
    return cell


def _setup_sheet(ws):
    ws.merge_cells("A1:G1")
    c           = ws["A1"]
    c.value     = "LAPORAN HITUNG BUAH"
    c.font      = Font(name="Arial", bold=True, size=14, color=HEADER_FG)
    c.fill      = PatternFill("solid", start_color=HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    for col, h in enumerate(["No", "Nama Klien", "Waktu Penyediaan",
                              "Jenis Buah", "Jumlah", "Tanggal", "Waktu"], 1):
        _cell(ws, 2, col, h, bold=True, bg=SUBHDR_BG, fg="1B5E20", align="center")
    ws.row_dimensions[2].height = 22

    for col, w in enumerate([6, 22, 20, 25, 10, 14, 10], 1):
        ws.column_dimensions[get_column_letter(col)].width = w


def save_to_excel(detections: List[FruitDetection], client: str,
                  meal_time: str, fruit_label: str,
                  excel_path: str = "hasil_hitung_buah.xlsx"):
    now    = datetime.now()
    date_  = now.strftime("%d/%m/%Y")
    time_  = now.strftime("%H:%M:%S")
    counts = {}
    for d in detections:
        counts[d.label] = counts.get(d.label, 0) + 1
    if not counts:
        counts = {fruit_label: 0}

    if os.path.exists(excel_path):
        wb        = load_workbook(excel_path)
        ws        = wb.active  # type: ignore
        assert ws is not None, "Worksheet tidak dapat diakses"
        start_row = ws.max_row + 1
        if ws["A1"].value != "LAPORAN HITUNG BUAH":
            ws.delete_rows(1, ws.max_row)
            _setup_sheet(ws)
            start_row = 3
    else:
        wb        = Workbook()
        ws        = wb.active  # type: ignore
        assert ws is not None, "Worksheet tidak dapat dibuat"
        ws.title  = "Data Buah"
        _setup_sheet(ws)
        start_row = 3

    last_no = 0
    for row in ws.iter_rows(min_row=3, max_col=1, values_only=True):
        if isinstance(row[0], int):
            last_no = row[0]

    row_idx = start_row
    for jenis, jumlah in sorted(counts.items()):
        last_no += 1
        bg = ROW_ODD if last_no % 2 == 1 else ROW_EVEN
        _cell(ws, row_idx, 1, last_no,   align="center", bg=bg)
        _cell(ws, row_idx, 2, client,    align="left",   bg=bg)
        _cell(ws, row_idx, 3, meal_time, align="center", bg=bg)
        _cell(ws, row_idx, 4, jenis,     align="left",   bg=bg)
        _cell(ws, row_idx, 5, jumlah,    align="center", bg=bg)
        _cell(ws, row_idx, 6, date_,     align="center", bg=bg)
        _cell(ws, row_idx, 7, time_,     align="center", bg=bg)
        row_idx += 1

    ws.merge_cells(f"A{row_idx}:D{row_idx}")
    lbl           = ws.cell(row=row_idx, column=1,
                            value=f"TOTAL - {client} ({meal_time} | {fruit_label})")
    lbl.font      = Font(name="Arial", bold=True, size=11, color="1B5E20")
    lbl.fill      = PatternFill("solid", start_color=TOTAL_BG)
    lbl.alignment = Alignment(horizontal="right", vertical="center")
    lbl.border    = _thin_border()
    _cell(ws, row_idx, 5, len(detections), bold=True, bg=TOTAL_BG, fg="1B5E20", align="center")
    for col in (6, 7):
        _cell(ws, row_idx, col, "", bg=TOTAL_BG)

    ws.freeze_panes = "A3"
    wb.save(excel_path)

    print(f"\n  Data disimpan ke: {excel_path}")
    print(f"  Klien  : {client}")
    print(f"  Waktu  : {meal_time} | {date_} {time_}")
    print(f"  Buah   : {fruit_label}")
    for jenis, jumlah in sorted(counts.items()):
        print(f"  {jenis:25s}: {jumlah} buah")
    print(f"  Total  : {len(detections)} buah")


# ══════════════════════════════════════════════════════════════════════════════
#  Mode Gambar & Kamera
# ══════════════════════════════════════════════════════════════════════════════

def process_image(path: str, client: str, meal_time: str, fruit_label: str,
                  selected_key: Optional[str], min_area: int,
                  save_path: Optional[str] = None, debug: bool = False,
                  excel_path: str = "hasil_hitung_buah.xlsx"):
    frame = cv2.imread(path)
    if frame is None:
        print(f"[ERROR] Tidak bisa membuka: {path}")
        sys.exit(1)
    h, w = frame.shape[:2]
    if max(h, w) > 1200:
        scale = 1200 / max(h, w)
        frame = cv2.resize(frame, (int(w * scale), int(h * scale)))

    detections = detect_fruits(frame, min_area, selected_key)
    result     = draw_results(frame, detections, client, meal_time, fruit_label)

    if save_path:
        cv2.imwrite(save_path, result)
    if debug:
        cv2.imshow("Debug Masks", _debug_panel(frame))
    try:
        save_to_excel(detections, client, meal_time, fruit_label, excel_path)
    except Exception as e:
        print(f"[ERROR] Gagal simpan Excel: {e}")

    cv2.imshow("Penghitung Buah", result)
    print("Tekan tombol apa saja untuk keluar...")
    cv2.waitKey(0)
    cv2.destroyAllWindows()


def process_camera(camera_id: int, client: str, meal_time: str,
                   fruit_label: str, selected_key: Optional[str],
                   min_area: int, debug: bool = False,
                   excel_path: str = "hasil_hitung_buah.xlsx"):
    cap = cv2.VideoCapture(camera_id, cv2.CAP_DSHOW)
    if not cap.isOpened():
        print(f"[ERROR] Tidak bisa membuka kamera ID: {camera_id}")
        sys.exit(1)
    cap.set(cv2.CAP_PROP_FRAME_WIDTH,  1280)
    cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 720)
    print("Kamera aktif.")
    print("  [q/ESC] keluar+simpan  |  [s] screenshot  |  [+/-] sensitivitas  |  [d] debug")

    shot_count = 0
    show_debug = debug

    while True:
        ret, frame = cap.read()
        if not ret:
            break
        frame      = cv2.flip(frame, 1)
        detections = detect_fruits(frame, min_area, selected_key)
        result     = draw_results(frame, detections, client, meal_time, fruit_label)

        fh, _ = result.shape[:2]
        cv2.putText(result, f"min_area={min_area}  (+/-)", (10, fh - 12),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.4, (180,180,180), 1, cv2.LINE_AA)
        cv2.imshow("Penghitung Buah (Kamera)", result)
        if show_debug:
            cv2.imshow("Debug Masks", _debug_panel(frame))

        key = cv2.waitKey(1) & 0xFF
        if key in (ord('q'), 27):
            print("\nMenyimpan ke Excel...")
            try:
                save_to_excel(detections, client, meal_time, fruit_label, excel_path)
            except Exception as e:
                print(f"[ERROR] Gagal simpan Excel: {e}")
            break
        elif key == ord('s'):
            shot_count += 1
            fname = f"screenshot_{shot_count:03d}.jpg"
            cv2.imwrite(fname, result)
            try:
                save_to_excel(detections, client, meal_time, fruit_label, excel_path)
                print(f"Screenshot {fname} tersimpan ({len(detections)} buah -> Excel)")
            except Exception as e:
                print(f"[ERROR] Gagal simpan Excel: {e}")
        elif key in (ord('+'), ord('=')):
            min_area = min(min_area + 200, 20000)
            print(f"  min_area -> {min_area}")
        elif key == ord('-'):
            min_area = max(min_area - 200, 200)
            print(f"  min_area -> {min_area}")
        elif key == ord('d'):
            show_debug = not show_debug

    cap.release()
    cv2.destroyAllWindows()
    print("Kamera ditutup.")


def _debug_panel(frame: np.ndarray) -> np.ndarray:
    hsv   = preprocess(frame)
    masks = []
    for k in FRUIT_COLORS:
        m = cv2.cvtColor(get_mask_for_color(hsv, k), cv2.COLOR_GRAY2BGR)
        cv2.putText(m, k, (5, 20), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0,220,220), 1)
        masks.append(cv2.resize(m, (160, 120)))
    row1 = np.hstack(masks[:3])
    pad  = [np.zeros_like(masks[0])] * (3 - len(masks[3:]))
    row2 = np.hstack(masks[3:] + pad)
    return np.vstack([row1, row2])


# ══════════════════════════════════════════════════════════════════════════════
#  Main
# ══════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description="Penghitung Buah -> Excel")
    parser.add_argument("input",      nargs="?", default=None)
    parser.add_argument("--save",     metavar="OUTPUT", default=None)
    parser.add_argument("--camera",   type=int, default=0)
    parser.add_argument("--min-area", type=int, default=1500)
    parser.add_argument("--excel",    default="hasil_hitung_buah.xlsx")
    parser.add_argument("--debug",    action="store_true")
    args = parser.parse_args()

    print("=" * 50)
    print("  Penghitung Buah - OpenCV + Excel")
    print("=" * 50)

    client, meal_time, selected_key = ask_session_info()
    fruit_label = (FRUIT_COLORS[selected_key][0][3]
                   if selected_key else "Semua Buah")

    if args.input:
        process_image(args.input, client, meal_time, fruit_label,
                      selected_key, args.min_area, args.save,
                      args.debug, args.excel)
    else:
        process_camera(args.camera, client, meal_time, fruit_label,
                       selected_key, args.min_area, args.debug, args.excel)


if __name__ == "__main__":
    main()
