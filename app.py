"""
Flask API untuk Fruit Counter
==============================
Menjalankan: python app.py
Akses: http://localhost:5000
"""

from flask import Flask, render_template, request, jsonify, send_file
from flask_cors import CORS
import cv2
import numpy as np
import base64
from io import BytesIO
from datetime import datetime
import os
import threading
import json

from fruit_counter import (
    FRUIT_COLORS, MEAL_TIMES, detect_fruits, draw_results,
    save_to_excel, preprocess, get_mask_for_color
)

app = Flask(__name__)
CORS(app)

# State global untuk kamera live
camera_state = {
    'active': False,
    'min_area': 1500,
    'detections': [],
    'debug_mode': False,
    'session': {
        'client': '',
        'meal_time': '',
        'fruit_label': ''
    }
}


@app.route('/')
def index():
    """Halaman utama"""
    return render_template('index.html')


@app.route('/api/config', methods=['GET'])
def get_config():
    """Dapatkan konfigurasi fruit colors dan meal times"""
    fruit_options = []
    for key, colors_list in FRUIT_COLORS.items():
        fruit_options.append({
            'key': key,
            'label': colors_list[0][3]
        })
    
    return jsonify({
        'fruits': fruit_options,
        'meals': MEAL_TIMES,
        'default_min_area': 1500
    })


@app.route('/api/session', methods=['POST'])
def set_session():
    """Set informasi sesi (klien, waktu, buah)"""
    data = request.json
    camera_state['session'] = {
        'client': data.get('client', ''),
        'meal_time': data.get('meal_time', ''),
        'fruit_label': data.get('fruit_label', '')
    }
    return jsonify({'status': 'ok'})


@app.route('/api/min-area', methods=['POST'])
def set_min_area():
    """Ubah sensitivitas minimum area"""
    data = request.json
    camera_state['min_area'] = data.get('min_area', 1500)
    return jsonify({'status': 'ok', 'min_area': camera_state['min_area']})


@app.route('/api/process-image', methods=['POST'])
def process_image_api():
    """Proses gambar yang di-upload"""
    if 'image' not in request.files:
        return jsonify({'error': 'Tidak ada file image'}), 400
    
    file = request.files['image']
    nparr = np.frombuffer(file.read(), np.uint8)
    frame = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
    
    if frame is None:
        return jsonify({'error': 'Tidak bisa membaca gambar'}), 400
    
    # Resize jika terlalu besar
    h, w = frame.shape[:2]
    if max(h, w) > 1200:
        scale = 1200 / max(h, w)
        frame = cv2.resize(frame, (int(w * scale), int(h * scale)))
    
    # Deteksi buah
    selected_key = request.form.get('fruit_key')
    min_area = int(request.form.get('min_area', 1500))
    detections = detect_fruits(frame, min_area, selected_key if selected_key != 'all' else None)
    
    # Draw hasil
    result = draw_results(
        frame, detections,
        camera_state['session']['client'],
        camera_state['session']['meal_time'],
        camera_state['session']['fruit_label']
    )
    
    # Convert ke base64
    _, buffer = cv2.imencode('.jpg', result)
    img_base64 = base64.b64encode(buffer).decode()
    
    # Hitung jumlah per jenis buah
    counts = {}
    for d in detections:
        counts[d.label] = counts.get(d.label, 0) + 1
    
    return jsonify({
        'image': f'data:image/jpeg;base64,{img_base64}',
        'total': len(detections),
        'counts': counts
    })


@app.route('/api/camera-frame', methods=['GET'])
def get_camera_frame():
    """Ambil frame dari kamera (untuk live preview)"""
    cap = cv2.VideoCapture(0, cv2.CAP_DSHOW)
    if not cap.isOpened():
        return jsonify({'error': 'Kamera tidak tersedia'}), 500
    
    ret, frame = cap.read()
    cap.release()
    
    if not ret:
        return jsonify({'error': 'Tidak bisa membaca frame'}), 500
    
    # Flip horizontal (mirror)
    frame = cv2.flip(frame, 1)
    
    # Deteksi
    selected_key = request.args.get('fruit_key')
    min_area = int(request.args.get('min_area', 1500))
    detections = detect_fruits(frame, min_area, selected_key if selected_key != 'all' else None)
    
    # Draw
    result = draw_results(
        frame, detections,
        camera_state['session']['client'],
        camera_state['session']['meal_time'],
        camera_state['session']['fruit_label']
    )
    
    # Convert
    _, buffer = cv2.imencode('.jpg', result)
    img_base64 = base64.b64encode(buffer).decode()
    
    counts = {}
    for d in detections:
        counts[d.label] = counts.get(d.label, 0) + 1
    
    return jsonify({
        'image': f'data:image/jpeg;base64,{img_base64}',
        'total': len(detections),
        'counts': counts
    })


@app.route('/api/save-excel', methods=['POST'])
def save_excel_api():
    """Simpan hasil ke Excel"""
    data = request.json
    
    # Buat detection objects dummy untuk save_to_excel
    from fruit_counter import FruitDetection
    
    client = data.get('client', 'Unknown')
    meal_time = data.get('meal_time', 'Unknown')
    fruit_label = data.get('fruit_label', 'Unknown')
    counts = data.get('counts', {})
    total = data.get('total', 0)
    
    # Buat fake detections berdasarkan counts
    detections = []
    for label, count in counts.items():
        for _ in range(count):
            detections.append(FruitDetection(
                x=0, y=0, w=0, h=0, radius=0,
                cx=0, cy=0, area=0,
                color_key='', label=label,
                display_color=(0, 0, 0)
            ))
    
    try:
        excel_path = data.get('excel_path', 'hasil_hitung_buah.xlsx')
        save_to_excel(detections, client, meal_time, fruit_label, excel_path)
        return jsonify({
            'status': 'ok',
            'message': f'Data disimpan ke {excel_path}',
            'file': excel_path
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/download-excel')
def download_excel():
    """Download file Excel"""
    excel_path = request.args.get('file', 'hasil_hitung_buah.xlsx')
    if os.path.exists(excel_path):
        return send_file(excel_path, as_attachment=True)
    return jsonify({'error': 'File tidak ditemukan'}), 404


@app.route('/api/excel-data', methods=['GET'])
def get_excel_data():
    """Baca dan preview data dari file Excel"""
    try:
        from openpyxl import load_workbook
        
        excel_path = request.args.get('file', 'hasil_hitung_buah.xlsx')
        
        if not os.path.exists(excel_path):
            return jsonify({'error': 'File Excel tidak ditemukan'}), 404
        
        wb = load_workbook(excel_path)
        ws = wb.active  # type: ignore
        assert ws is not None, "Worksheet tidak dapat diakses"
        
        # Baca semua data
        rows = []
        for row in ws.iter_rows(min_row=3, values_only=True):
            if row[0] is None:
                break
            rows.append({
                'no': row[0],
                'client': row[1],
                'meal_time': row[2],
                'fruit': row[3],
                'jumlah': row[4],
                'tanggal': row[5],
                'waktu': row[6]
            })
        
        return jsonify({
            'status': 'ok',
            'file': excel_path,
            'rows': rows,
            'total_rows': len(rows)
        })
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/open-excel', methods=['GET'])
def open_excel():
    """Buka file Excel di aplikasi default sistem"""
    try:
        import subprocess
        import platform
        
        excel_path = request.args.get('file', 'hasil_hitung_buah.xlsx')
        
        if not os.path.exists(excel_path):
            return jsonify({'error': 'File Excel tidak ditemukan'}), 404
        
        # Dapatkan path absolut
        abs_path = os.path.abspath(excel_path)
        
        # Buka file sesuai OS
        system = platform.system()
        if system == 'Windows':
            os.startfile(abs_path)
        elif system == 'Darwin':  # macOS
            subprocess.Popen(['open', abs_path])
        else:  # Linux
            subprocess.Popen(['xdg-open', abs_path])
        
        return jsonify({
            'status': 'ok',
            'message': f'File dibuka: {abs_path}'
        })
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500


if __name__ == '__main__':
    print("\n" + "="*50)
    print("  Fruit Counter - Web UI")
    print("="*50)
    print("  Buka: http://localhost:5000")
    print("  Tekan: Ctrl+C untuk keluar")
    print("="*50 + "\n")
    
    app.run(debug=True, host='0.0.0.0', port=5000)
